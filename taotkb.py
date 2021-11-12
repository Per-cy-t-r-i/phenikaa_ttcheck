from openpyxl import Workbook, load_workbook
import calendar
from datetime import datetime

SRC = 'tkbtheoweb.xlsx'
TKB = 'tkb.xlsm'

wb_source = load_workbook(SRC)
tt_web = wb_source.active

wb_tkb = load_workbook(TKB, keep_vba = True)
tkb_xl = wb_tkb.active

today = datetime.today().strftime('%d-%m-%y')
cdate = int(today[0:2])
cmonth = int(today[3:5])
cyear = int(today[6:8])

thu_theo_cell= {
    '2' : 'C',
    '3' : 'D',
    '4' : 'E',
    '5' : 'F',
    '6' : 'G',
    '7' : 'H'
}
tiet_theo_cell= {
    '1':'2',
    '2':'3',
    '3':'4',
    '4':'5',
    '5':'6',
    '6':'8',
    '7':'9',
    '8':'10',
    '9':'11',
    '10':'12',
    '11':'14',
    '12':'15',
}

def tach(string):
    ans = ''
    if len(string)==1:
        return string
    a = int(string[0])
    b = int(string[1:])
    for i in range(a,b+1):
        ans+= str(i)+" "
    ans = ans.split(" ")
    ans.pop()
    return ans

def Thu_tiet(data):
    thu = []
    tiets = []
    for i in range(len(data)):
        if data[i] =='(':
            tiet = ''
            thu.append(data[i-1])
            j = i
            while data[j] != ')':
                tiet += data[j]
                j+=1
            tiet = tiet.replace("(T", "")
            tiet = tiet.replace("-", "")
            tiets.append(tach(tiet))
    return thu, tiets
    
def get_cells(thu, tiet):
    cells = []
    for i in range(len(thu)):
        cell = thu_theo_cell[thu[i]]
        for j in tiet[i]:
            cells.append(cell+tiet_theo_cell[j])
    return cells
    
def reset_tkb():
    chr ='CDEFGH'
    num =['2','3','4','5','6','8','9','10','11','12','14','15']
    for i in chr:
        for j in num:
            tkb_xl[i+j]= ""
            
def den_luc(date_month_year):
    date = int(date_month_year[0:2])
    month = int(date_month_year[3:5])
    year = int(date_month_year[6:8])
    if year >cyear:
        return False
    if month>cmonth:    
        return False
    if date >cdate and month == cmonth:
        return False    
    return True
    
    
def chua_het(date_month_year):   
    date = int(date_month_year[0:2])
    month = int(date_month_year[3:5])
    year = int(date_month_year[6:8])
    if year <cyear:
        return False
    if month <cmonth:
        return False
    if date<cdate and month ==cmonth:
        return False
    return True
    
    
def Fill_tkb():
    i = '2'
    while tt_web['C'+i].value != None:   
        ten_mon = tt_web['C'+i].value
        data = tt_web['H'+i].value
        date_month_year = data[:8]
        ex_date_month_year = data[9:17]
        if den_luc(date_month_year) and chua_het(ex_date_month_year):
            thu, tiet = Thu_tiet(data)
            cells = get_cells(thu, tiet)
            for cell in cells:
                tkb_xl[cell] = ten_mon       
        i = str(int(i)+1)
        
reset_tkb()
Fill_tkb()       



wb_tkb.save(TKB)
wb_tkb.close()
wb_source.close()


  

    
